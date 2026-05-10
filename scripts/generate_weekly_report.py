#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import os
import re
import subprocess
import sys
import tempfile
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import column_index_from_string, get_column_letter
    from PIL import Image as PILImage
except ModuleNotFoundError as exc:
    missing = exc.name or "dependency"
    raise SystemExit(
        f"缺少 Python 依赖：{missing}。\n"
        "请先运行：pip install -r requirements.txt\n"
        "或使用 Codex bundled Python 运行本脚本。"
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
CONFIG_DIR = ROOT / "config"


@dataclass
class ReportRecord:
    record_id: str
    brand: str
    product_name: str
    launch_date: date | None
    price: str
    category: str
    series: str
    selling_point: str
    ingredients: str
    image_urls: list[str]
    remark: str


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_configs() -> dict[str, dict[str, Any]]:
    configs = {
        "dingtalk": load_json(CONFIG_DIR / "dingtalk.json"),
        "field_mapping": load_json(CONFIG_DIR / "field_mapping.json"),
        "report_rules": load_json(CONFIG_DIR / "report_rules.json"),
        "excel_layout": load_json(CONFIG_DIR / "excel_layout.json"),
        "font_files": load_json(CONFIG_DIR / "font_files.json"),
    }
    return configs


def validate_config(configs: dict[str, dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    dingtalk = configs["dingtalk"]
    if not dingtalk.get("baseId"):
        warnings.append("config/dingtalk.json 缺少 baseId")
    if not dingtalk.get("tableId"):
        warnings.append("config/dingtalk.json 缺少 tableId")
    if not dingtalk.get("streamableHttpUrl") and not dingtalk.get("serverName"):
        warnings.append("config/dingtalk.json 需要 streamableHttpUrl 或 serverName")

    fields = configs["field_mapping"].get("standardFields", {})
    for key in ("brand", "productName", "launchDate"):
        if key not in fields:
            warnings.append(f"field_mapping.json 缺少必需标准字段：{key}")

    rules = configs["report_rules"]
    if not rules.get("defaultBrands"):
        warnings.append("report_rules.json 缺少 defaultBrands")
    if not rules.get("categoryRule", {}).get("priority"):
        warnings.append("report_rules.json 缺少 categoryRule.priority")

    layout = configs["excel_layout"]
    if not layout.get("worksheetName"):
        warnings.append("excel_layout.json 缺少 worksheetName")
    image_cm = layout.get("image", {}).get("heightCm")
    if not isinstance(image_cm, (int, float)) or image_cm <= 0:
        warnings.append("excel_layout.json 的 image.heightCm 必须为正数")

    font_files = configs["font_files"]
    for font_key in ("chineseFont", "latinFont"):
        font_cfg = font_files.get(font_key, {})
        candidates = [font_cfg.get("sourcePath"), ROOT / str(font_cfg.get("projectPath", ""))]
        if not any(candidate and Path(candidate).exists() for candidate in candidates):
            warnings.append(f"font_files.json 中 {font_key} 的字体文件当前不可访问；Excel 仍会写入字体名称")
    return warnings


def explain_config(configs: dict[str, dict[str, Any]]) -> str:
    dingtalk = configs["dingtalk"]
    fields = configs["field_mapping"]["standardFields"]
    rules = configs["report_rules"]
    layout = configs["excel_layout"]
    font_files = configs["font_files"]
    lines = [
        "当前配置摘要：",
        f"- 钉钉连接：baseId={dingtalk.get('baseId')}，tableId={dingtalk.get('tableId')}，"
        f"连接方式={'streamableHttpUrl' if dingtalk.get('streamableHttpUrl') else dingtalk.get('serverName')}",
        f"- 默认周期：最近一个完整周六到周五",
        f"- 默认品牌：{'、'.join(rules.get('defaultBrands', []))}",
        f"- 品类规则：{' > '.join(rules.get('categoryRule', {}).get('priority', []))} > 主品类",
        f"- 备注规则：{'，'.join(rules.get('remarkRule', {}).get('order', []))}",
        f"- 图片高度：{layout.get('image', {}).get('heightCm')}cm",
        f"- 中文字体：{font_files.get('chineseFont', {}).get('excelName')}",
        f"- 英文字体：{font_files.get('latinFont', {}).get('excelName')}",
        "- 标准字段：",
    ]
    for key, cfg in fields.items():
        names = " / ".join(cfg.get("fieldNames", []))
        lines.append(f"  - {key}: {cfg.get('label')} ({names})")
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从钉钉 AI 表生成竞品新品周报 Excel")
    parser.add_argument("--start-date", help="开始日期，格式 YYYY-MM-DD")
    parser.add_argument("--end-date", help="结束日期，格式 YYYY-MM-DD")
    parser.add_argument("--brands", help="品牌范围，使用英文逗号分隔，例如：霸王茶姬,古茗")
    parser.add_argument("--output", help="输出 xlsx 路径")
    parser.add_argument("--validate-config", action="store_true", help="校验配置后退出")
    parser.add_argument("--explain-config", action="store_true", help="解释当前配置后退出")
    return parser.parse_args()


def resolve_date_window(args: argparse.Namespace) -> tuple[date, date]:
    if args.start_date or args.end_date:
        if not args.start_date or not args.end_date:
            raise SystemExit("--start-date 和 --end-date 必须同时提供")
        start = date.fromisoformat(args.start_date)
        end = date.fromisoformat(args.end_date)
        if start > end:
            raise SystemExit("--start-date 不能晚于 --end-date")
        return start, end

    today = date.today()
    friday_weekday = 4
    days_since_friday = (today.weekday() - friday_weekday) % 7
    end = today - timedelta(days=days_since_friday)
    start = end - timedelta(days=6)
    return start, end


def parse_brands(arg: str | None, rules: dict[str, Any]) -> list[str]:
    if not arg:
        return list(rules.get("defaultBrands", []))
    brands = [part.strip() for part in arg.split(",") if part.strip()]
    if not brands:
        raise SystemExit("--brands 不能为空")
    return brands


def mcporter_selector(dingtalk_cfg: dict[str, Any], tool_name: str) -> list[str]:
    url = dingtalk_cfg.get("streamableHttpUrl", "").strip()
    if url:
        return ["mcporter", "call", url, f".{tool_name}"]
    server = dingtalk_cfg.get("serverName", "dingtalk-ai-table")
    return ["mcporter", "call", f"{server}.{tool_name}"]


def call_dingtalk_tool(dingtalk_cfg: dict[str, Any], tool_name: str, payload: dict[str, Any]) -> dict[str, Any]:
    cmd = mcporter_selector(dingtalk_cfg, tool_name) + ["--args", json.dumps(payload, ensure_ascii=False)]
    try:
        result = subprocess.run(cmd, check=False, capture_output=True, text=True)
    except FileNotFoundError as exc:
        raise SystemExit("找不到 mcporter，请先安装并配置钉钉 AI 表 MCP。") from exc

    if result.returncode != 0:
        raise SystemExit(f"mcporter 调用失败：{result.stderr.strip() or result.stdout.strip()}")
    try:
        data = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise SystemExit(f"mcporter 返回不是 JSON：{result.stdout[:1000]}") from exc
    if data.get("status") not in (None, "success", "ok"):
        raise SystemExit(f"钉钉工具返回失败：{json.dumps(data.get('error', data), ensure_ascii=False)}")
    return data


def fetch_table_fields(configs: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    dingtalk = configs["dingtalk"]
    payload = {"baseId": dingtalk["baseId"], "tableIds": [dingtalk["tableId"]]}
    data = call_dingtalk_tool(dingtalk, "get_tables", payload)
    tables = data.get("data", {}).get("tables", [])
    if not tables:
        raise SystemExit("没有读取到钉钉表结构，请检查 baseId/tableId。")
    return tables[0].get("fields", [])


def resolve_field_ids(field_mapping: dict[str, Any], table_fields: list[dict[str, Any]]) -> dict[str, str]:
    by_id = {f.get("fieldId"): f for f in table_fields if f.get("fieldId")}
    by_name = {f.get("fieldName"): f for f in table_fields if f.get("fieldName")}
    resolved: dict[str, str] = {}
    missing_required: list[str] = []

    for key, cfg in field_mapping.get("standardFields", {}).items():
        field_id = cfg.get("fieldId")
        if field_id and field_id in by_id:
            resolved[key] = field_id
            continue
        for name in cfg.get("fieldNames", []):
            if name in by_name:
                resolved[key] = by_name[name]["fieldId"]
                break
        if key not in resolved and cfg.get("required"):
            missing_required.append(f"{key} ({cfg.get('label')})")

    if missing_required:
        raise SystemExit("缺少必需钉钉字段：" + "、".join(missing_required))
    return resolved


def query_records(
    configs: dict[str, dict[str, Any]],
    field_ids: dict[str, str],
    start: date,
    end: date,
) -> list[dict[str, Any]]:
    dingtalk = configs["dingtalk"]
    date_field = field_ids.get("launchDate")
    if not date_field:
        raise SystemExit("字段映射中缺少 launchDate")

    records: list[dict[str, Any]] = []
    cursor = None
    seen_cursors: set[str] = set()
    requested_field_ids = list(dict.fromkeys(field_ids.values()))

    while True:
        payload: dict[str, Any] = {
            "baseId": dingtalk["baseId"],
            "tableId": dingtalk["tableId"],
            "limit": 100,
            "fieldIds": requested_field_ids,
            "filters": {
                "operator": "and",
                "operands": [
                    {"operator": "not_before", "operands": [date_field, start.isoformat()]},
                    {"operator": "not_after", "operands": [date_field, end.isoformat()]},
                ],
            },
            "sort": [{"fieldId": date_field, "direction": "asc"}],
        }
        if cursor:
            payload["cursor"] = cursor
        data = call_dingtalk_tool(dingtalk, "query_records", payload)
        batch = data.get("data", {}).get("records", []) or []
        records.extend(batch)
        next_cursor = data.get("data", {}).get("nextCursor")
        if not next_cursor or next_cursor in seen_cursors or not batch:
            break
        seen_cursors.add(next_cursor)
        cursor = next_cursor
    return records


def extract_cell(raw: Any, extract: str) -> Any:
    if raw is None:
        return "" if extract != "multi_select_names" and extract != "attachment_image_urls" else []
    if extract == "select_name":
        if isinstance(raw, dict):
            return str(raw.get("name", "")).strip()
        return str(raw).strip()
    if extract == "multi_select_names":
        if isinstance(raw, list):
            names = []
            for item in raw:
                if isinstance(item, dict):
                    name = str(item.get("name", "")).strip()
                    if name:
                        names.append(name)
                elif item:
                    names.append(str(item).strip())
            return names
        if isinstance(raw, dict):
            return [str(raw.get("name", "")).strip()] if raw.get("name") else []
        return [str(raw).strip()] if raw else []
    if extract == "rich_text_markdown":
        if isinstance(raw, dict):
            text = raw.get("markdown") or raw.get("text") or ""
        else:
            text = str(raw)
        return clean_text(text)
    if extract == "attachment_image_urls":
        if not isinstance(raw, list):
            return []
        urls = []
        for item in raw:
            if isinstance(item, dict) and item.get("type") == "image" and item.get("url"):
                urls.append(item["url"])
        return urls
    if extract == "date":
        return parse_date_value(raw)
    if extract == "generic":
        if isinstance(raw, dict):
            return raw.get("name") or raw.get("text") or raw.get("value") or ""
        if isinstance(raw, list):
            return [extract_cell(item, "generic") for item in raw]
        return raw
    return clean_text(str(raw))


def parse_date_value(raw: Any) -> date | None:
    if not raw:
        return None
    if isinstance(raw, (int, float)):
        # Dingtalk date fields normally return ISO strings; keep a safe Excel-like fallback.
        return datetime.fromtimestamp(raw / 1000).date()
    text = str(raw).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return None


def clean_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n")]
    lines = [line for line in lines if line]
    return "\n".join(lines).strip()


def normalize_records(
    raw_records: list[dict[str, Any]],
    configs: dict[str, dict[str, Any]],
    field_ids: dict[str, str],
    brands: list[str],
) -> list[ReportRecord]:
    fields_cfg = configs["field_mapping"]["standardFields"]
    rules = configs["report_rules"]
    brand_set = set(brands)
    normalized: list[ReportRecord] = []

    for raw_record in raw_records:
        cells = raw_record.get("cells", {})
        values: dict[str, Any] = {}
        for standard_key, field_id in field_ids.items():
            cfg = fields_cfg.get(standard_key, {})
            values[standard_key] = extract_cell(cells.get(field_id), cfg.get("extract", "text"))

        brand = str(values.get("brand") or "").strip()
        product_name = str(values.get("productName") or "").strip()
        if not brand or not product_name or brand not in brand_set:
            continue

        category = compute_category(values, rules)
        remark = compute_remark(values, rules)
        normalized.append(
            ReportRecord(
                record_id=raw_record.get("recordId", ""),
                brand=brand,
                product_name=product_name,
                launch_date=values.get("launchDate"),
                price=str(values.get("price") or ""),
                category=category,
                series=str(values.get("series") or ""),
                selling_point=str(values.get("sellingPoint") or ""),
                ingredients=str(values.get("ingredients") or ""),
                image_urls=values.get("appearanceImages") or [],
                remark=remark,
            )
        )

    brand_order = {brand: i for i, brand in enumerate(brands)}
    normalized.sort(
        key=lambda r: (
            brand_order.get(r.brand, 9999),
            r.launch_date or date.min,
            r.product_name,
        )
    )
    return normalized


def compute_category(values: dict[str, Any], rules: dict[str, Any]) -> str:
    category_rule = rules.get("categoryRule", {})
    additional = values.get(category_rule.get("additionalField", "additionalCategory")) or []
    if isinstance(additional, str):
        additional = [additional]
    for candidate in category_rule.get("priority", []):
        if candidate in additional:
            return candidate
    main = values.get(category_rule.get("mainField", "mainCategory"))
    if isinstance(main, list):
        return str(main[0]) if main else ""
    return str(main or "")


def truthy_return_status(value: Any, rule: dict[str, Any]) -> bool:
    if isinstance(value, list):
        return any(truthy_return_status(item, rule) for item in value)
    if isinstance(value, bool):
        return value
    text = str(value or "").strip()
    if not text:
        return False
    ignored = set(rule.get("ignoredValues", []))
    truthy = set(rule.get("truthyValues", []))
    if text in ignored:
        return False
    return text in truthy or rule.get("showValue", "回归") in text


def compute_remark(values: dict[str, Any], rules: dict[str, Any]) -> str:
    remark_rule = rules.get("remarkRule", {})
    parts: list[str] = []
    for key in remark_rule.get("order", []):
        if key == "returnStatus":
            if truthy_return_status(values.get(key), remark_rule.get("returnStatus", {})):
                parts.append(remark_rule.get("returnStatus", {}).get("showValue", "回归"))
        elif key == "collaboration":
            text = clean_text(str(values.get(key) or ""))
            if text:
                suffix = remark_rule.get("collaboration", {}).get("suffix", "联名")
                parts.append(text if text.endswith(suffix) else text + suffix)
        else:
            text = clean_text(str(values.get(key) or ""))
            if text:
                parts.append(text)
    return remark_rule.get("separator", "，").join(parts) if parts else remark_rule.get("emptyText", "/")


def format_title_date(start: date, end: date, template: str) -> str:
    return template.format(start_m=start.month, start_d=start.day, end_m=end.month, end_d=end.day)


def cm_to_points(cm: float) -> float:
    return cm / 2.54 * 72


def cm_to_pixels(cm: float) -> int:
    return int(round(cm / 2.54 * 96))


def estimate_row_height(text: str, col_chars: float, min_height: float, max_height: float) -> float:
    if not text:
        return min_height
    lines = 0
    for part in str(text).split("\n"):
        visual_len = sum(2 if ord(ch) > 127 else 1 for ch in part)
        lines += max(1, math.ceil(visual_len / max(col_chars * 1.8, 1)))
    return max(min_height, min(max_height, lines * 15.5))


def price_with_slash_wrap(text: str) -> str:
    text = clean_text(text)
    return text.replace("/", "/\n")


def build_price_rich_text(text: str, font_name: str, size: int, strike_pattern: str) -> CellRichText | str:
    if not text:
        return ""
    normal = InlineFont(rFont=font_name, sz=size)
    strike = InlineFont(rFont=font_name, sz=size, strike=True)
    pattern = re.compile(r"\(([^()]*)\)")
    price_pattern = re.compile(strike_pattern)
    rich = CellRichText()
    pos = 0
    found = False
    for match in pattern.finditer(text):
        if match.start() > pos:
            rich.append(TextBlock(normal, text[pos : match.start()]))
        inner = match.group(1)
        if price_pattern.search(inner):
            rich.append(TextBlock(normal, "("))
            rich.append(TextBlock(strike, inner))
            rich.append(TextBlock(normal, ")"))
            found = True
        else:
            rich.append(TextBlock(normal, match.group(0)))
        pos = match.end()
    if pos < len(text):
        rich.append(TextBlock(normal, text[pos:]))
    return rich if found else text


def set_cell_price(cell, text: str, configs: dict[str, dict[str, Any]], wrap_after_slash: bool = False) -> None:
    rules = configs["report_rules"].get("priceRule", {})
    layout = configs["excel_layout"]
    font_name = layout.get("fonts", {}).get("latinExcelName") or layout.get("fonts", {}).get("chineseExcelName")
    size = layout.get("fonts", {}).get("defaultSize", 10)
    value = price_with_slash_wrap(text) if wrap_after_slash else clean_text(text)
    if rules.get("strikeParenthesizedPrice", True):
        cell.value = build_price_rich_text(value, font_name, size, rules.get("parenthesizedPricePattern", r"\d+(?:\.\d+)?\s*元"))
    else:
        cell.value = value


def apply_base_style(cell, font_name: str, size: int = 10, bold: bool = False, fill: str | None = None, border: Border | None = None):
    cell.font = Font(name=font_name, size=size, bold=bold, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if border:
        cell.border = border


def merge_and_style(ws, range_ref: str, value: Any, font_name: str, size: int, bold: bool, fill: str | None, border: Border | None):
    ws.merge_cells(range_ref)
    cell = ws[range_ref.split(":")[0]]
    cell.value = value
    apply_base_style(cell, font_name, size=size, bold=bold, fill=fill, border=border)
    if border:
        for row in ws[range_ref]:
            for c in row:
                c.border = border


def download_image(url: str, record_id: str, cache_dir: Path) -> Path | None:
    if not url:
        return None
    cache_dir.mkdir(parents=True, exist_ok=True)
    suffix = ".png"
    clean_id = re.sub(r"[^A-Za-z0-9_-]+", "_", record_id or "image")
    target = cache_dir / f"{clean_id}{suffix}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            target.write_bytes(resp.read())
        return target
    except Exception as exc:
        print(f"WARNING: 图片下载失败：{record_id} {exc}", file=sys.stderr)
        return None


def add_image(ws, image_path: Path, anchor: str, height_cm: float, max_width_px: int) -> None:
    try:
        with PILImage.open(image_path) as pil:
            width, height = pil.size
        target_height = cm_to_pixels(height_cm)
        target_width = int(width * target_height / height) if height else target_height
        if target_width > max_width_px:
            target_width = max_width_px
            target_height = int(height * target_width / width) if width else cm_to_pixels(height_cm)
        image = XLImage(str(image_path))
        image.height = target_height
        image.width = target_width
        image.anchor = anchor
        ws.add_image(image)
    except Exception as exc:
        print(f"WARNING: 图片插入失败：{image_path} {exc}", file=sys.stderr)


def build_workbook(
    records: list[ReportRecord],
    configs: dict[str, dict[str, Any]],
    start: date,
    end: date,
    brands: list[str],
    output_path: Path,
) -> None:
    layout = configs["excel_layout"]
    rules = configs["report_rules"]
    colors = layout["colors"]
    fonts = layout["fonts"]
    heights = layout["rowHeights"]
    chinese_font = fonts["chineseExcelName"]
    latin_font = fonts["latinExcelName"]
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = layout["worksheetName"]
    ws.sheet_view.showGridLines = False

    for col, width in layout["columns"].items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[2].height = heights["topSpacer"]
    merge_and_style(ws, "B3:H3", format_title_date(start, end, layout["titleTemplate"]), chinese_font, fonts["titleSize"], True, colors["white"], None)
    ws.row_dimensions[3].height = heights["title"]
    merge_and_style(ws, "B5:H5", layout["introText"], chinese_font, fonts["introSize"], False, colors["white"], None)
    ws["B5"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[5].height = heights["intro"]

    header_row = layout["summary"]["headerRow"]
    summary_cols = layout["summary"]["columns"]
    headers = layout["summary"]["headers"]
    for key in ("brand", "count", "category", "productName", "launchDate", "price", "remark"):
        col = summary_cols[key]
        cell = ws[f"{col}{header_row}"]
        cell.value = "" if key == "brand" else headers[key]
        apply_base_style(cell, chinese_font, bold=True, fill=colors["headerFill"], border=border)
    ws.row_dimensions[header_row].height = heights["summaryHeader"]

    grouped: dict[str, list[ReportRecord]] = defaultdict(list)
    for record in records:
        grouped[record.brand].append(record)

    row = layout["summary"]["startRow"]
    displayed_brands = [brand for brand in brands if grouped.get(brand)]
    for brand in displayed_brands:
        brand_records = grouped[brand]
        start_row = row
        end_row = row + len(brand_records) - 1
        if start_row != end_row:
            ws.merge_cells(f"B{start_row}:B{end_row}")
            ws.merge_cells(f"C{start_row}:C{end_row}")
        ws[f"B{start_row}"] = brand
        ws[f"C{start_row}"] = len(brand_records)
        for cell_ref in (f"B{start_row}", f"C{start_row}"):
            apply_base_style(ws[cell_ref], chinese_font if cell_ref.startswith("B") else latin_font, bold=True, fill=colors["brandFill"] if cell_ref.startswith("B") else None, border=border)
        for rr in range(start_row, end_row + 1):
            for col in ("B", "C"):
                ws[f"{col}{rr}"].border = border

        for record in brand_records:
            ws[f"D{row}"] = record.category
            ws[f"E{row}"] = record.product_name
            ws[f"F{row}"] = record.launch_date
            if record.launch_date:
                ws[f"F{row}"].number_format = "m/d"
            set_cell_price(ws[f"G{row}"], record.price, configs, wrap_after_slash=True)
            ws[f"H{row}"] = record.remark
            for col in ("D", "E", "F", "G", "H"):
                apply_base_style(ws[f"{col}{row}"], chinese_font if col in ("D", "E", "H") else latin_font, border=border)
            ws.row_dimensions[row].height = heights["summaryWrapped"] if "\n" in str(ws[f"G{row}"].value) else heights["summaryDefault"]
            row += 1

    note_row = row + 1
    merge_and_style(
        ws,
        f"B{note_row}:H{note_row}",
        layout["trackedBrandsPrefix"] + "、".join(brands),
        chinese_font,
        fonts["defaultSize"],
        False,
        colors["white"],
        None,
    )
    ws[f"B{note_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[note_row].height = heights["summaryWrapped"]

    detail_row = note_row + layout["details"]["blankRowsAfterSummary"] + 1
    image_cache = output_path.parent / "_image_cache"
    for brand in displayed_brands:
        brand_records = grouped[brand]
        merge_and_style(
            ws,
            f"B{detail_row}:H{detail_row}",
            layout["details"]["brandHeaderTemplate"].format(brand=brand, count=len(brand_records)),
            chinese_font,
            fonts["defaultSize"],
            True,
            colors["detailHeaderFill"],
            border,
        )
        ws.row_dimensions[detail_row].height = heights["detailHeader"]
        detail_row += 1

        for record in brand_records:
            row_values = {
                "productName": record.product_name,
                "series": record.series,
                "sellingPoint": record.selling_point,
                "price": record.price,
                "ingredients": record.ingredients,
                "appearanceImages": "",
            }
            for key, label in layout["details"]["rowsPerProduct"]:
                ws[f"B{detail_row}"] = label
                apply_base_style(ws[f"B{detail_row}"], chinese_font, bold=True, fill=colors["detailLabelFill"], border=border)
                value_range = f"C{detail_row}:H{detail_row}"
                ws.merge_cells(value_range)
                value_cell = ws[f"C{detail_row}"]
                if key == "price":
                    set_cell_price(value_cell, record.price, configs, wrap_after_slash=False)
                else:
                    value_cell.value = row_values.get(key, "")
                apply_base_style(value_cell, chinese_font, bold=key == "productName", fill=colors["detailLabelFill"] if key == "productName" else None, border=border)
                value_cell.alignment = Alignment(horizontal="left" if key in ("sellingPoint", "ingredients") else "center", vertical="center", wrap_text=True)
                for merged_row in ws[value_range]:
                    for cell in merged_row:
                        cell.border = border
                        if key == "productName":
                            cell.fill = PatternFill("solid", fgColor=colors["detailLabelFill"])

                if key == "appearanceImages":
                    ws.row_dimensions[detail_row].height = cm_to_points(layout["image"]["heightCm"])
                    if record.image_urls:
                        image_path = download_image(record.image_urls[0], record.record_id, image_cache)
                        if image_path:
                            add_image(
                                ws,
                                image_path,
                                f"{layout['image']['anchorColumn']}{detail_row}",
                                layout["image"]["heightCm"],
                                layout["image"].get("maxWidthPx", 520),
                            )
                elif key in ("sellingPoint", "ingredients"):
                    text = str(row_values.get(key, ""))
                    ws.row_dimensions[detail_row].height = estimate_row_height(
                        text,
                        70,
                        heights["detailTextMin"],
                        heights["detailTextMax"],
                    )
                else:
                    ws.row_dimensions[detail_row].height = heights["detailDefault"]
                detail_row += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def default_output_path(start: date, end: date, layout: dict[str, Any]) -> Path:
    output_dir = ROOT / layout.get("outputDirectory", "outputs")
    return output_dir / f"竞品新品周报{start.isoformat()}_{end.isoformat()}.xlsx"


def main() -> int:
    args = parse_args()
    configs = load_configs()

    warnings = validate_config(configs)
    if args.validate_config:
        if warnings:
            print("配置校验完成，有警告：")
            for warning in warnings:
                print(f"- {warning}")
        else:
            print("配置校验通过。")
        return 0 if not [w for w in warnings if "缺少" in w or "必须" in w or "需要" in w] else 1

    if args.explain_config:
        print(explain_config(configs))
        if warnings:
            print("\n配置警告：")
            for warning in warnings:
                print(f"- {warning}")
        return 0

    start, end = resolve_date_window(args)
    brands = parse_brands(args.brands, configs["report_rules"])
    table_fields = fetch_table_fields(configs)
    field_ids = resolve_field_ids(configs["field_mapping"], table_fields)
    raw_records = query_records(configs, field_ids, start, end)
    records = normalize_records(raw_records, configs, field_ids, brands)
    output_path = Path(args.output) if args.output else default_output_path(start, end, configs["excel_layout"])
    if not output_path.is_absolute():
        output_path = ROOT / output_path
    build_workbook(records, configs, start, end, brands, output_path)
    print(f"已生成：{output_path}")
    print(f"记录数：{len(records)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
